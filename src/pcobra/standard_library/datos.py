"""Utilidades tabulares sin depender de :mod:`pandas`.

El proyecto original utilizaba ``pandas`` para todas las operaciones sobre
conjuntos de datos. Durante las pruebas del repositorio las dependencias
pesadas no siempre están disponibles, por lo que este módulo implementa las
funciones clave únicamente con la biblioteca estándar y el stub ligero de
:nmod:`numpy` incluido en la base de código. Las funciones mantienen la misma
firma pública para que el código existente continúe funcionando.
"""

from __future__ import annotations

import builtins
import importlib.util
import csv
import json
import math
import statistics
from collections import Counter, defaultdict
from contextlib import suppress
from itertools import groupby
from pathlib import Path
from typing import Any, Callable, Iterable, Iterator, Mapping, MutableMapping, Sequence

import numpy as np

Registro = dict[str, Any]
Tabla = list[Registro]

__all__ = [
    "leer_csv",
    "leer_json",
    "escribir_csv",
    "escribir_json",
    "leer_excel",
    "escribir_excel",
    "leer_parquet",
    "escribir_parquet",
    "leer_feather",
    "escribir_feather",
    "describir",
    "correlacion_pearson",
    "correlacion_spearman",
    "matriz_covarianza",
    "calcular_percentiles",
    "resumen_rapido",
    "seleccionar_columnas",
    "filtrar",
    "mutar_columna",
    "separar_columna",
    "unir_columnas",
    "agrupar_y_resumir",
    "tabla_cruzada",
    "pivotar_ancho",
    "pivotar_largo",
    "ordenar_tabla",
    "combinar_tablas",
    "rellenar_nulos",
    "desplegar_tabla",
    "pivotar_tabla",
    "a_listas",
    "de_listas",
]


# ---------------------------------------------------------------------------
# Utilidades internas
# ---------------------------------------------------------------------------


def _es_numero(valor: Any) -> bool:
    return isinstance(valor, (int, float)) and not isinstance(valor, bool)


def _convertir_valor_crudo(valor: str | None) -> Any:
    if valor is None:
        return None
    texto = valor.strip()
    if texto == "":
        return None
    with suppress(ValueError):
        return int(texto)
    with suppress(ValueError):
        return float(texto)
    texto_lower = texto.lower()
    if texto_lower in {"true", "false"}:
        return texto_lower == "true"
    return texto


def _copiar_tabla(tabla: Tabla) -> Tabla:
    return [dict(fila) for fila in tabla]


def _materializar_tabla(
    datos: Iterable[Registro]
    | Mapping[str, Sequence[Any]]
    | "DataFrame"  # type: ignore[name-defined]
) -> Tabla:
    if isinstance(datos, list):
        return [dict(fila) for fila in datos]
    if isinstance(datos, Mapping):
        columnas = {clave: list(valor) for clave, valor in datos.items()}
        longitudes = {len(valores) for valores in columnas.values()}
        if len(longitudes) > 1:
            raise ValueError("Todas las columnas deben tener la misma longitud")
        longitud = longitudes.pop() if columnas else 0
        return [
            {columna: columnas[columna][indice] for columna in columnas}
            for indice in range(longitud)
        ]
    to_dict = getattr(datos, "to_dict", None)
    if callable(to_dict):
        try:
            registros = to_dict(orient="records")  # type: ignore[call-arg]
        except TypeError:
            pass
        else:
            if isinstance(registros, list):
                return [dict(fila) for fila in registros]
    if isinstance(datos, Iterable):
        resultado: Tabla = []
        for fila in datos:
            if isinstance(fila, Mapping):
                resultado.append(dict(fila))
            else:
                raise TypeError("Las filas deben ser diccionarios")
        return resultado
    raise TypeError("Formato de datos no soportado")


def _modulo_disponible(nombre: str) -> bool:
    """Comprueba si un módulo opcional está disponible."""

    return importlib.util.find_spec(nombre) is not None


def _columnas(tabla: Tabla) -> list[str]:
    columnas: list[str] = []
    for fila in tabla:
        for columna in fila.keys():
            if columna not in columnas:
                columnas.append(columna)
    return columnas


def _columnas_numericas(tabla: Tabla) -> list[str]:
    return [
        columna
        for columna in _columnas(tabla)
        if any(_es_numero(fila.get(columna)) for fila in tabla)
    ]


def _valores_columna(
    tabla: Tabla, columna: str, *, solo_numeros: bool = False
) -> list[Any]:
    valores: list[Any] = []
    for fila in tabla:
        valor = fila.get(columna)
        if solo_numeros and not _es_numero(valor):
            continue
        valores.append(valor)
    return valores


def _pares_columnas(tabla: Tabla, columna_a: str, columna_b: str) -> list[tuple[float, float]]:
    pares: list[tuple[float, float]] = []
    for fila in tabla:
        valor_a = fila.get(columna_a)
        valor_b = fila.get(columna_b)
        if _es_numero(valor_a) and _es_numero(valor_b):
            pares.append((float(valor_a), float(valor_b)))
    return pares


def _valores_numericos(tabla: Tabla, columna: str) -> list[float]:
    return [float(valor) for valor in _valores_columna(tabla, columna, solo_numeros=True)]


def _valores_numericos_alineados(tabla: Tabla, columnas: Sequence[str]) -> list[list[float]]:
    alineados: list[list[float]] = []
    for fila in tabla:
        fila_convertida: list[float] = []
        for columna in columnas:
            valor = fila.get(columna)
            if not _es_numero(valor):
                break
            fila_convertida.append(float(valor))
        else:
            alineados.append(fila_convertida)
    return alineados


def _aplicar_agregacion(nombre: str, valores: Sequence[Any]) -> Any:
    filtrados = [valor for valor in valores if valor is not None]
    if not filtrados:
        return None
    if nombre in {"sum", "suma"}:
        total = 0
        for valor in filtrados:
            total += valor  # type: ignore[operator]
        return total
    if nombre in {"mean", "promedio", "avg"}:
        numeros = [float(valor) for valor in filtrados]
        return sum(numeros) / len(numeros)
    if nombre in {"max", "min"}:
        comparables = list(filtrados)
        return getattr(builtins, nombre)(comparables)  # type: ignore[arg-type]
    if nombre in {"count", "conteo", "len"}:
        return len(filtrados)
    raise ValueError(f"Agregación desconocida: {nombre}")


def _calcular_covarianza(valores_a: Sequence[float], valores_b: Sequence[float]) -> float:
    n = len(valores_a)
    if n < 2:
        return 0.0
    media_a = sum(valores_a) / n
    media_b = sum(valores_b) / n
    acumulado = sum((a - media_a) * (b - media_b) for a, b in zip(valores_a, valores_b))
    return acumulado / (n - 1)


def _calcular_pearson(valores_a: Sequence[float], valores_b: Sequence[float]) -> float:
    covarianza = _calcular_covarianza(valores_a, valores_b)
    desviacion_a = math.sqrt(_calcular_covarianza(valores_a, valores_a))
    desviacion_b = math.sqrt(_calcular_covarianza(valores_b, valores_b))
    if desviacion_a == 0 or desviacion_b == 0:
        return 0.0
    return covarianza / (desviacion_a * desviacion_b)


def _rango_promedio(valores: Sequence[float]) -> list[float]:
    pares = sorted((valor, indice) for indice, valor in enumerate(valores))
    rangos = [0.0] * len(valores)
    posicion = 0
    while posicion < len(pares):
        valor, indice = pares[posicion]
        j = posicion
        while j < len(pares) and pares[j][0] == valor:
            j += 1
        rango_promedio = sum(range(posicion + 1, j + 1)) / (j - posicion)
        for k in range(posicion, j):
            rangos[pares[k][1]] = rango_promedio
        posicion = j
    return rangos


# ---------------------------------------------------------------------------
# Lectura y escritura
# ---------------------------------------------------------------------------


def leer_csv(
    ruta: str | Path,
    *,
    separador: str = ",",
    encoding: str = "utf-8",
    limite_filas: int | None = None,
) -> Tabla:
    try:
        with Path(ruta).open("r", encoding=encoding, newline="") as archivo:
            lector = csv.DictReader(archivo, delimiter=separador, strict=True)
            resultado: Tabla = []
            columnas = lector.fieldnames or []
            for indice, fila in enumerate(lector):
                if limite_filas is not None and indice >= limite_filas:
                    break
                if None in fila or len(fila) != len(columnas):
                    raise ValueError("No fue posible leer el CSV: formato inválido")
                resultado.append({clave: _convertir_valor_crudo(valor) for clave, valor in fila.items()})
            return resultado
    except (csv.Error, OSError, UnicodeDecodeError) as exc:
        raise ValueError(f"No fue posible leer el CSV: {exc}") from exc


def escribir_csv(
    datos: Iterable[Registro] | Mapping[str, Sequence[Any]] | "DataFrame",
    ruta: str | Path,
    *,
    separador: str = ",",
    encoding: str = "utf-8",
    aniadir: bool = False,
    incluir_indice: bool = False,
) -> None:
    tabla = _materializar_tabla(datos)
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)

    modo = "a" if aniadir and ruta.exists() else "w"
    escribir_encabezado = modo == "w" or (modo == "a" and ruta.stat().st_size == 0)

    with ruta.open(modo, encoding=encoding, newline="") as archivo:
        columnas = _columnas(tabla)
        escritor = csv.DictWriter(archivo, fieldnames=columnas, delimiter=separador)
        if escribir_encabezado:
            escritor.writeheader()
        for indice, fila in enumerate(tabla):
            fila_a_escribir = dict(fila)
            if incluir_indice:
                fila_a_escribir.setdefault("index", indice)
            for clave, valor in fila_a_escribir.items():
                if valor is None:
                    fila_a_escribir[clave] = ""
            escritor.writerow(fila_a_escribir)


def leer_json(ruta: str | Path, *, orient: str | None = None, lineas: bool = False) -> Tabla:
    try:
        texto = Path(ruta).read_text(encoding="utf-8")
    except OSError as exc:
        raise ValueError(f"No fue posible leer el JSON: {exc}") from exc

    try:
        if lineas:
            registros = [json.loads(linea) for linea in texto.splitlines() if linea.strip()]
        else:
            datos = json.loads(texto)
            if isinstance(datos, list):
                registros = datos
            elif isinstance(datos, dict):
                registros = datos.get("records") if orient == "records" else [datos]
            else:
                raise TypeError("Estructura JSON no soportada")
    except (json.JSONDecodeError, TypeError) as exc:
        raise ValueError(f"No fue posible leer el JSON: {exc}") from exc
    return _materializar_tabla(registros)


def escribir_json(
    datos: Iterable[Registro] | Mapping[str, Sequence[Any]] | "DataFrame",
    ruta: str | Path,
    *,
    indent: int | None = 2,
    lineas: bool = False,
    aniadir: bool = False,
) -> None:
    tabla = _materializar_tabla(datos)
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)

    if lineas:
        modo = "a" if aniadir else "w"
        claves_existentes: set[str] = set()
        if aniadir and ruta.exists():
            for linea in ruta.read_text(encoding="utf-8").splitlines():
                linea = linea.strip()
                if linea:
                    try:
                        claves_existentes.update(json.loads(linea).keys())
                    except json.JSONDecodeError:
                        continue
        claves_totales = set(claves_existentes)
        for fila in tabla:
            claves_totales.update(fila.keys())
        with ruta.open(modo, encoding="utf-8") as archivo:
            for fila in tabla:
                completo = {clave: fila.get(clave) for clave in claves_totales}
                archivo.write(json.dumps(completo, ensure_ascii=False))
                archivo.write("\n")
        return

    texto = json.dumps(tabla, ensure_ascii=False, indent=indent)
    ruta.write_text(texto, encoding="utf-8")


def _asegurar_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ModuleNotFoundError as exc:  # pragma: no cover - depende del entorno
        raise ValueError("Para trabajar con Excel instala 'openpyxl'.") from exc
    return openpyxl


def escribir_excel(
    datos: Iterable[Registro] | Mapping[str, Sequence[Any]] | "DataFrame",
    ruta: str | Path,
    *,
    hoja: str = "Hoja1",
    engine: str | None = None,
) -> None:
    _asegurar_openpyxl()
    tabla = _materializar_tabla(datos)
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)

    from openpyxl import Workbook  # type: ignore[import-not-found]

    libro = Workbook()
    hoja_obj = libro.active
    hoja_obj.title = hoja

    columnas = _columnas(tabla)
    if columnas:
        hoja_obj.append(columnas)
    for fila in tabla:
        hoja_obj.append([fila.get(columna) for columna in columnas])

    libro.save(ruta)


def leer_excel(
    ruta: str | Path,
    *,
    hoja: str | int | None = 0,
    encabezado: int | None = 0,
    engine: str | None = None,
) -> Tabla:
    _asegurar_openpyxl()
    from openpyxl import load_workbook  # type: ignore[import-not-found]

    libro = load_workbook(Path(ruta), read_only=True, data_only=True)
    if hoja is None:
        hoja_obj = libro.active
    elif isinstance(hoja, int):
        hoja_obj = libro.worksheets[hoja]
    else:
        hoja_obj = libro[hoja]

    filas = list(hoja_obj.iter_rows(values_only=True))
    if not filas:
        return []

    if encabezado is None:
        columnas = [i for i in range(len(filas[0]))]
        inicio_datos = 0
    else:
        encabezados = filas[encabezado]
        columnas = [enc if enc is not None else indice for indice, enc in enumerate(encabezados)]
        inicio_datos = encabezado + 1

    resultado: Tabla = []
    for fila in filas[inicio_datos:]:
        registro: Registro = {}
        for indice, columna in enumerate(columnas):
            registro[columna] = fila[indice] if indice < len(fila) else None
        resultado.append(registro)
    return resultado


def _asegurar_pyarrow(accion: str) -> tuple[Any, Any]:
    if not _modulo_disponible("pyarrow"):
        mensaje = "Instala 'pyarrow' o 'fastparquet'." if "Parquet" in accion else "instala el paquete opcional 'pyarrow'."
        raise ValueError(f"No fue posible {accion}: {mensaje}")
    try:
        import pyarrow as pa  # type: ignore[import-not-found]
        import pyarrow.feather as feather  # type: ignore[import-not-found]
        import pyarrow.parquet as pq  # type: ignore[import-not-found]
    except ModuleNotFoundError as exc:  # pragma: no cover - depende del entorno
        mensaje = "Instala 'pyarrow' o 'fastparquet'." if "Parquet" in accion else "instala el paquete opcional 'pyarrow'."
        raise ValueError(f"No fue posible {accion}: {mensaje}") from exc
    return pa, (feather, pq)


def escribir_parquet(
    datos: Iterable[Registro] | Mapping[str, Sequence[Any]] | "DataFrame",
    ruta: str | Path,
    *,
    engine: str | None = None,
) -> None:
    pa, (_, pq) = _asegurar_pyarrow("escribir el archivo Parquet")
    tabla = _materializar_tabla(datos)
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    pq.write_table(pa.Table.from_pylist(tabla), ruta)


def leer_parquet(
    ruta: str | Path,
    *,
    engine: str | None = None,
) -> Tabla:
    pa, (_, pq) = _asegurar_pyarrow("leer el archivo Parquet")
    tabla = pq.read_table(Path(ruta))
    return [dict(zip(tabla.column_names, fila)) for fila in zip(*[col.to_pylist() for col in tabla.columns])]


def escribir_feather(
    datos: Iterable[Registro] | Mapping[str, Sequence[Any]] | "DataFrame",
    ruta: str | Path,
    *,
    compression: str | None = None,
) -> None:
    pa, (feather, _) = _asegurar_pyarrow("escribir el archivo Feather")
    tabla = _materializar_tabla(datos)
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    feather.write_feather(pa.Table.from_pylist(tabla), ruta, compression=compression)


def leer_feather(ruta: str | Path) -> Tabla:
    pa, (feather, _) = _asegurar_pyarrow("leer el archivo Feather")
    tabla = feather.read_table(Path(ruta))
    return [dict(zip(tabla.column_names, fila)) for fila in zip(*[col.to_pylist() for col in tabla.columns])]


# ---------------------------------------------------------------------------
# Transformaciones
# ---------------------------------------------------------------------------


def a_listas(tabla: Iterable[Registro]) -> dict[str, list[Any]]:
    filas = _materializar_tabla(tabla)
    columnas = _columnas(filas)
    resultado: dict[str, list[Any]] = {columna: [] for columna in columnas}
    for fila in filas:
        for columna in columnas:
            resultado[columna].append(fila.get(columna))
    return resultado


def de_listas(columnas: Mapping[str, Sequence[Any]]) -> Tabla:
    longitudes = {len(valores) for valores in columnas.values()}
    if len(longitudes) > 1:
        raise ValueError("Todas las columnas deben tener la misma longitud")
    longitud = longitudes.pop() if columnas else 0
    claves = list(columnas.keys())
    return [
        {clave: columnas[clave][indice] for clave in claves}
        for indice in range(longitud)
    ]


def seleccionar_columnas(tabla: Iterable[Registro], columnas: Sequence[str]) -> Tabla:
    filas = _materializar_tabla(tabla)
    resultado: Tabla = []
    for fila in filas:
        seleccionado: Registro = {}
        for columna in columnas:
            if columna not in fila:
                raise KeyError(columna)
            seleccionado[columna] = fila[columna]
        resultado.append(seleccionado)
    return resultado


def filtrar(tabla: Iterable[Registro], condicion: Callable[[Registro], bool]) -> Tabla:
    filas = _materializar_tabla(tabla)
    resultado: Tabla = []
    for fila in filas:
        try:
            if condicion(fila):
                resultado.append(dict(fila))
        except Exception as exc:  # pragma: no cover - errores usuario
            raise ValueError(f"La condición de filtrado falló: {exc}") from exc
    return resultado


def mutar_columna(
    tabla: Iterable[Registro],
    columna: str,
    transformacion: Callable[[Registro], Any],
    *,
    crear_si_no_existe: bool = True,
) -> Tabla:
    filas = _materializar_tabla(tabla)
    resultado: Tabla = []
    for fila in filas:
        nuevo = dict(fila)
        if not crear_si_no_existe and columna not in nuevo:
            raise KeyError(columna)
        nuevo[columna] = transformacion(nuevo)
        resultado.append(nuevo)
    return resultado


def separar_columna(
    tabla: Iterable[Registro],
    columna: str,
    *,
    en: Sequence[str],
    separador: str = ",",
    eliminar_original: bool = True,
    relleno: Any = "",
    descartar_nulos: bool = False,
) -> Tabla:
    filas = _materializar_tabla(tabla)
    resultado: Tabla = []
    for fila in filas:
        valor = fila.get(columna)
        if valor is None:
            if descartar_nulos:
                continue
            partes = [relleno for _ in en]
        else:
            partes = str(valor).split(separador)
            if len(partes) < len(en):
                partes.extend(relleno for _ in range(len(en) - len(partes)))
            if len(partes) > len(en):
                partes = partes[: len(en)]
        nuevo = dict(fila)
        if eliminar_original:
            nuevo.pop(columna, None)
        for nombre, parte in zip(en, partes):
            nuevo[nombre] = parte
        resultado.append(nuevo)
    return resultado


def unir_columnas(
    tabla: Iterable[Registro],
    columnas: Sequence[str],
    destino: str,
    *,
    separador: str = " ",
    omitir_nulos: bool = False,
    relleno: str = "sin dato",
    eliminar_original: bool = True,
) -> Tabla:
    filas = _materializar_tabla(tabla)
    resultado: Tabla = []
    for fila in filas:
        partes: list[str] = []
        for columna in columnas:
            valor = fila.get(columna)
            if valor is None:
                if omitir_nulos:
                    continue
                partes.append(relleno)
            else:
                partes.append(str(valor))
        nuevo = dict(fila)
        if omitir_nulos and not partes:
            nuevo[destino] = None
        else:
            nuevo[destino] = separador.join(partes)
        if eliminar_original:
            for columna in columnas:
                nuevo.pop(columna, None)
        resultado.append(nuevo)
    return resultado


def agrupar_y_resumir(
    tabla: Iterable[Registro],
    *,
    por: Sequence[str],
    agregaciones: Mapping[str, str | Sequence[str]],
) -> Tabla:
    filas = _materializar_tabla(tabla)
    grupos: dict[tuple[Any, ...], dict[str, list[Any]]] = {}
    orden_claves: list[tuple[Any, ...]] = []
    for fila in filas:
        clave = tuple(fila.get(columna) for columna in por)
        if clave not in grupos:
            grupos[clave] = {columna: [] for columna in agregaciones}
            orden_claves.append(clave)
        grupo = grupos[clave]
        for columna in agregaciones:
            grupo.setdefault(columna, []).append(fila.get(columna))

    resultado: Tabla = []
    for clave in orden_claves:
        fila_salida: Registro = {col: val for col, val in zip(por, clave)}
        datos_grupo = grupos[clave]
        for columna, agg in agregaciones.items():
            valores = datos_grupo.get(columna, [])
            if isinstance(agg, Sequence) and not isinstance(agg, str):
                for nombre in agg:
                    fila_salida[f"{columna}_{nombre}"] = _aplicar_agregacion(nombre, valores)
            else:
                nombre = str(agg)
                fila_salida[f"{columna}_{nombre}"] = _aplicar_agregacion(nombre, valores)
        resultado.append(fila_salida)
    return resultado


def tabla_cruzada(
    tabla: Iterable[Registro],
    filas: str,
    columnas: str,
    *,
    valores: str | None = None,
    aggfunc: str | Callable[[Sequence[Any]], Any] = "count",
    normalizar: str | None = None,
) -> Tabla:
    datos = _materializar_tabla(tabla)
    orden_filas: list[Any] = []
    orden_columnas: list[Any] = []
    acumulados: dict[Any, dict[Any, list[Any]]] = {}

    for fila in datos:
        clave_fila = fila.get(filas)
        clave_columna = fila.get(columnas)
        if clave_fila not in orden_filas:
            orden_filas.append(clave_fila)
        if clave_columna not in orden_columnas:
            orden_columnas.append(clave_columna)
        acumulados.setdefault(clave_fila, {}).setdefault(clave_columna, []).append(
            fila.get(valores) if valores is not None else 1
        )

    resultado: Tabla = []
    for clave_fila in orden_filas:
        fila_salida: Registro = {filas: clave_fila}
        for clave_columna in orden_columnas:
            valores_celda = acumulados.get(clave_fila, {}).get(clave_columna, [])
            if valores is None:
                valor = len(valores_celda)
            else:
                if callable(aggfunc):
                    valor = aggfunc(valores_celda)
                else:
                    valor = _aplicar_agregacion(str(aggfunc), valores_celda)
                if valor is None:
                    valor = 0
            fila_salida[clave_columna] = valor
        resultado.append(fila_salida)

    if normalizar == "filas":
        for fila in resultado:
            total = sum(valor for clave, valor in fila.items() if clave != filas)
            if total:
                for clave in orden_columnas:
                    fila[clave] = fila[clave] / total
    elif normalizar == "columnas":
        totales = Counter({clave: 0.0 for clave in orden_columnas})
        for fila in resultado:
            for clave in orden_columnas:
                totales[clave] += fila[clave]
        for fila in resultado:
            for clave in orden_columnas:
                total = totales[clave]
                fila[clave] = fila[clave] / total if total else 0.0

    return resultado


def pivotar_ancho(
    tabla: Iterable[Registro],
    *,
    id_columnas: str,
    nombres_desde: str,
    valores_desde: str,
    valores_relleno: Any = None,
) -> Tabla:
    filas = _materializar_tabla(tabla)
    indices: list[Any] = []
    columnas_dinamicas: list[Any] = []
    resultado: dict[Any, Registro] = {}

    for fila in filas:
        clave = fila.get(id_columnas)
        if clave not in indices:
            indices.append(clave)
            resultado[clave] = {id_columnas: clave}
        nombre = fila.get(nombres_desde)
        if nombre not in columnas_dinamicas:
            columnas_dinamicas.append(nombre)
        valor = fila.get(valores_desde, valores_relleno)
        if valor is None:
            valor = valores_relleno
        resultado[clave][f"{valores_desde}_{nombre}"] = valor

    salida: Tabla = []
    for clave in indices:
        fila = resultado[clave]
        for nombre in columnas_dinamicas:
            columna = f"{valores_desde}_{nombre}"
            fila.setdefault(columna, valores_relleno)
        salida.append(dict(fila))
    return salida


def pivotar_largo(
    tabla: Iterable[Registro],
    *,
    columnas: Sequence[str],
    id_columnas: str,
    nombres_a: str,
    valores_a: str,
    eliminar_nulos: bool = False,
) -> Tabla:
    filas = _materializar_tabla(tabla)
    resultado: Tabla = []
    for fila in filas:
        identificador = fila.get(id_columnas)
        for columna in columnas:
            valor = fila.get(columna)
            if valor is None and eliminar_nulos:
                continue
            resultado.append({
                id_columnas: identificador,
                nombres_a: columna,
                valores_a: valor,
            })
    return resultado


def ordenar_tabla(
    tabla: Iterable[Registro],
    *,
    por: Sequence[str],
    ascendente: bool | Sequence[bool] = True,
) -> Tabla:
    filas = _materializar_tabla(tabla)
    if isinstance(ascendente, bool):
        ordenes = [ascendente] * len(por)
    else:
        ordenes = list(ascendente)
    if len(ordenes) != len(por):
        raise ValueError("La lista de orden debe coincidir con las columnas")

    resultado = _copiar_tabla(filas)
    for columna, asc in reversed(list(zip(por, ordenes))):
        resultado.sort(key=lambda fila: (fila.get(columna) is None, fila.get(columna)), reverse=not asc)
    return resultado


def combinar_tablas(
    izquierda: Iterable[Registro],
    derecha: Iterable[Registro],
    *,
    claves: Mapping[str, str] | Sequence[str] | tuple[str, str],
    tipo: str = "inner",
) -> Tabla:
    tabla_izquierda = _materializar_tabla(izquierda)
    tabla_derecha = _materializar_tabla(derecha)

    if isinstance(claves, Mapping):
        clave_izquierda = claves.get("izquierda")
        clave_derecha = claves.get("derecha")
    elif isinstance(claves, Sequence) and len(claves) == 2:
        clave_izquierda, clave_derecha = claves
    else:
        raise ValueError("Debes proporcionar las claves de unión")

    indice_derecho: dict[Any, list[Registro]] = defaultdict(list)
    for fila in tabla_derecha:
        indice_derecho[fila.get(clave_derecha)].append(fila)

    resultado: Tabla = []
    coincidencias_derecha: Counter[Any] = Counter()

    for fila_izquierda in tabla_izquierda:
        clave = fila_izquierda.get(clave_izquierda)
        coincidencias = indice_derecho.get(clave)
        if coincidencias:
            for fila_derecha in coincidencias:
                combinada = dict(fila_izquierda)
                for columna, valor in fila_derecha.items():
                    if columna in combinada and combinada[columna] != valor:
                        combinada[f"{columna}_derecha"] = valor
                    else:
                        combinada[columna] = valor
                resultado.append(combinada)
            coincidencias_derecha[clave] += len(coincidencias)
        elif tipo in {"left", "outer"}:
            combinada = dict(fila_izquierda)
            for columna in _columnas(tabla_derecha):
                if columna not in combinada:
                    combinada[columna] = None
            resultado.append(combinada)

    if tipo == "outer":
        for clave, filas_derecha in indice_derecho.items():
            if coincidencias_derecha[clave]:
                continue
            for fila_derecha in filas_derecha:
                combinada = {columna: None for columna in _columnas(tabla_izquierda)}
                combinada.update(fila_derecha)
                resultado.append(combinada)

    return resultado


def rellenar_nulos(
    tabla: Iterable[Registro],
    reemplazos: Mapping[str, Any] | Any,
) -> Tabla:
    filas = _materializar_tabla(tabla)
    if isinstance(reemplazos, Mapping):
        reemplazos_map = reemplazos
    else:
        reemplazos_map = defaultdict(lambda: reemplazos)
    resultado: Tabla = []
    for fila in filas:
        nuevo = dict(fila)
        for columna, valor in list(nuevo.items()):
            if valor is None and columna in reemplazos_map:
                nuevo[columna] = reemplazos_map[columna]
        resultado.append(nuevo)
    return resultado


def desplegar_tabla(
    tabla: Iterable[Registro],
    *,
    identificadores: Sequence[str] | str,
    valores: Sequence[str] | None = None,
    var_name: str = "variable",
    value_name: str = "value",
) -> Tabla:
    filas = _materializar_tabla(tabla)
    ids = [identificadores] if isinstance(identificadores, str) else list(identificadores)
    columnas = valores if valores is not None else [col for col in _columnas(filas) if col not in ids]

    resultado: Tabla = []
    for columna in columnas:
        for fila in filas:
            base = {col: fila.get(col) for col in ids}
            registro = dict(base)
            registro[var_name] = columna
            registro[value_name] = fila.get(columna)
            resultado.append(registro)
    return resultado


def pivotar_tabla(
    tabla: Iterable[Registro],
    *,
    index: str | Sequence[str],
    columnas: str,
    valores: str | Sequence[str],
    agregacion: Mapping[str, str | Sequence[str]] | str = "mean",
) -> Tabla:
    filas = _materializar_tabla(tabla)
    indices = [index] if isinstance(index, str) else list(index)
    valores_lista = [valores] if isinstance(valores, str) else list(valores)

    if isinstance(agregacion, Mapping):
        agregaciones = agregacion
    else:
        agregaciones = {valor: agregacion for valor in valores_lista}

    orden_indices: list[tuple[Any, ...]] = []
    orden_columnas: list[Any] = []
    acumulados: dict[tuple[Any, ...], dict[Any, dict[str, list[Any]]]] = {}

    for fila in filas:
        clave_indice = tuple(fila.get(col) for col in indices)
        clave_columna = fila.get(columnas)
        if clave_indice not in acumulados:
            acumulados[clave_indice] = {}
            orden_indices.append(clave_indice)
        if clave_columna not in orden_columnas:
            orden_columnas.append(clave_columna)
        destino = acumulados[clave_indice].setdefault(clave_columna, {valor: [] for valor in valores_lista})
        for valor in valores_lista:
            destino.setdefault(valor, []).append(fila.get(valor))

    resultado: Tabla = []
    for clave_indice in orden_indices:
        fila_salida: Registro = {col: val for col, val in zip(indices, clave_indice)}
        for clave_columna in orden_columnas:
            valores_columna = acumulados[clave_indice].get(clave_columna, {})
            for valor in valores_lista:
                agreg = agregaciones.get(valor, "mean")
                valores_celda = valores_columna.get(valor, [])
                if isinstance(agreg, Sequence) and not isinstance(agreg, str):
                    for nombre in agreg:
                        fila_salida[f"{valor}_{nombre}_{clave_columna}"] = _aplicar_agregacion(nombre, valores_celda)
                else:
                    nombre = str(agreg)
                    fila_salida[f"{valor}_{nombre}_{clave_columna}"] = _aplicar_agregacion(nombre, valores_celda)
        resultado.append(fila_salida)
    return resultado


# ---------------------------------------------------------------------------
# Métricas estadísticas
# ---------------------------------------------------------------------------


def describir(datos: Iterable[Registro]) -> dict[str, dict[str, float]]:
    tabla = _materializar_tabla(datos)
    resultado: dict[str, dict[str, float]] = {}
    for columna in _columnas_numericas(tabla):
        valores = [float(valor) for valor in _valores_columna(tabla, columna, solo_numeros=True)]
        if not valores:
            continue
        media = sum(valores) / len(valores)
        if len(valores) > 1:
            varianza = sum((valor - media) ** 2 for valor in valores) / (len(valores) - 1)
            desviacion = math.sqrt(varianza)
        else:
            desviacion = 0.0
        resultado[columna] = {
            "count": float(len(valores)),
            "mean": media,
            "std": desviacion,
            "min": min(valores),
            "25%": float(np.percentile(valores, 25)),
            "50%": float(np.percentile(valores, 50)),
            "75%": float(np.percentile(valores, 75)),
            "max": max(valores),
        }
    return resultado


def correlacion_pearson(
    datos: Iterable[Registro],
    *,
    columnas: Sequence[str] | None = None,
) -> dict[str, dict[str, float]]:
    tabla = _materializar_tabla(datos)
    columnas_objetivo = list(columnas) if columnas is not None else _columnas_numericas(tabla)
    if not columnas_objetivo:
        raise ValueError("No hay columnas numéricas para calcular la correlación")

    disponibles = {col for fila in tabla for col in fila.keys()}
    for columna in columnas_objetivo:
        if columna not in disponibles:
            raise KeyError(columna)

    resultado: dict[str, dict[str, float]] = {}
    algun_valor = False
    for columna_a in columnas_objetivo:
        resultado[columna_a] = {}
        for columna_b in columnas_objetivo:
            pares = _pares_columnas(tabla, columna_a, columna_b)
            if not pares:
                valor = 0.0
            else:
                valores_a = [par[0] for par in pares]
                valores_b = [par[1] for par in pares]
                valor = _calcular_pearson(valores_a, valores_b)
                algun_valor = True
            if columna_a == columna_b and pares:
                valor = 1.0
            resultado[columna_a][columna_b] = valor
    if not algun_valor:
        raise ValueError("No hay suficientes datos numéricos para correlación")
    return resultado


def correlacion_spearman(
    datos: Iterable[Registro],
    *,
    columnas: Sequence[str] | None = None,
) -> dict[str, dict[str, float]]:
    tabla = _materializar_tabla(datos)
    columnas_objetivo = list(columnas) if columnas is not None else _columnas_numericas(tabla)
    if not columnas_objetivo:
        raise ValueError("No hay columnas numéricas para calcular la correlación")

    resultado: dict[str, dict[str, float]] = {}
    algun_valor = False
    for columna_a in columnas_objetivo:
        resultado[columna_a] = {}
        for columna_b in columnas_objetivo:
            pares = _pares_columnas(tabla, columna_a, columna_b)
            if not pares:
                valor = 0.0
            else:
                valores_a = [par[0] for par in pares]
                valores_b = [par[1] for par in pares]
                rangos_a = _rango_promedio(valores_a)
                rangos_b = _rango_promedio(valores_b)
                valor = _calcular_pearson(rangos_a, rangos_b)
                algun_valor = True
            if columna_a == columna_b and pares:
                valor = 1.0
            resultado[columna_a][columna_b] = valor
    if not algun_valor:
        raise ValueError("No hay suficientes datos numéricos para correlación")
    return resultado


def matriz_covarianza(
    datos: Iterable[Registro],
    *,
    columnas: Sequence[str] | None = None,
) -> dict[str, dict[str, float]]:
    tabla = _materializar_tabla(datos)
    columnas_objetivo = list(columnas) if columnas is not None else _columnas_numericas(tabla)
    if not columnas_objetivo:
        raise ValueError("No hay columnas numéricas para calcular la covarianza")

    resultado: dict[str, dict[str, float]] = {}
    algun_valor = False
    for columna_a in columnas_objetivo:
        resultado[columna_a] = {}
        for columna_b in columnas_objetivo:
            pares = _pares_columnas(tabla, columna_a, columna_b)
            if not pares:
                valores_a = _valores_numericos(tabla, columna_a)
                valores_b = _valores_numericos(tabla, columna_b)
                if not valores_a or not valores_b:
                    valor = 0.0
                else:
                    valor = _calcular_covarianza(valores_a, valores_b)
                    algun_valor = True
            else:
                valores_a = [par[0] for par in pares]
                valores_b = [par[1] for par in pares]
                valor = _calcular_covarianza(valores_a, valores_b)
                algun_valor = True
            resultado[columna_a][columna_b] = valor
    if not algun_valor:
        raise ValueError("No hay suficientes datos numéricos para covarianza")
    return resultado


def calcular_percentiles(
    datos: Iterable[Registro],
    *,
    columnas: Sequence[str] | None = None,
    percentiles: Sequence[int] = (25, 50, 75),
) -> dict[str, dict[str, float]]:
    if not percentiles:
        raise ValueError("Debes proporcionar al menos un percentil")
    if any(percentil < 0 or percentil > 100 for percentil in percentiles):
        raise ValueError("Los percentiles deben estar entre 0 y 100")

    tabla = _materializar_tabla(datos)
    columnas_objetivo = list(columnas) if columnas is not None else _columnas_numericas(tabla)
    resultado: dict[str, dict[str, float]] = {}
    for columna in columnas_objetivo:
        valores = _valores_numericos(tabla, columna)
        if not valores:
            continue
        resultado[columna] = {}
        for percentil in percentiles:
            resultado[columna][f"p{percentil}"] = float(np.percentile(valores, percentil))
    return resultado


def resumen_rapido(datos: Iterable[Registro]) -> list[dict[str, Any]]:
    tabla = _materializar_tabla(datos)
    resumen: list[dict[str, Any]] = []
    for columna in _columnas(tabla):
        valores = [fila.get(columna) for fila in tabla]
        no_nulos = [valor for valor in valores if valor is not None]
        registro: dict[str, Any] = {
            "columna": columna,
            "conteo": len(valores),
            "nulos": len(valores) - len(no_nulos),
        }
        if no_nulos:
            ejemplo = no_nulos[0]
            if hasattr(ejemplo, "isoformat"):
                registro["ejemplo"] = ejemplo.isoformat()
            else:
                registro["ejemplo"] = ejemplo
        else:
            registro["ejemplo"] = None

        if no_nulos and all(_es_numero(valor) for valor in no_nulos):
            numeros = [float(valor) for valor in no_nulos]
            registro["media"] = sum(numeros) / len(numeros)
            registro["min"] = min(numeros)
            registro["max"] = max(numeros)
        elif no_nulos and all(hasattr(valor, "isoformat") for valor in no_nulos):
            ordenadas = sorted(no_nulos)
            registro["min"] = ordenadas[0].isoformat()
            registro["max"] = ordenadas[-1].isoformat()
        elif no_nulos:
            registro["min"] = str(min(no_nulos))
        resumen.append(registro)
    return resumen
